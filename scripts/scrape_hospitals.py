#!/usr/bin/env python3
"""Mini ETL pipeline for hospitals.co.zw.

This script ingests facility data from multiple lightweight sources (raw files and
stub scrapers), normalises fields into a common shape, deduplicates near-matches
with fuzzy logic, infers missing attributes, and writes the canonical
``data/hospitals.json`` consumed by the frontend.

The helpers are intentionally small and rule-based so they can be upgraded with
ML/embeddings later without changing their public signatures.
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import pathlib
import re
from difflib import SequenceMatcher
from typing import Dict, Iterable, List, Optional, Tuple

ROOT = pathlib.Path(__file__).resolve().parents[1]
SCRAPED_OUTPUT = ROOT / "data" / "hospitals_scraped_new.json"
RAW_DIR = ROOT / "data" / "raw"
TODAY = dt.date.today().isoformat()
REMOTE_RAW_SOURCES = {
  "alliance_providers_pdf": {
    "url": "https://www.alliancehealth.co.zw/sites/default/files/attachments/Service%20Provider%20List%202020.pdf",
    "filename": "alliance_provider_list_2020.pdf",
  },
  "mcaz_premises_html": {
    "url": "https://onlineservices.mcaz.co.zw/onlineregister/frmPremisesRegister.aspx",
    "filename": "mcaz_premises_register.html",
  },
}
TRUSTED_SOURCES = {
  "hpa_registered_facilities",
  "mcaz_pharmacies_2024",
  "mohcc_official",
}

Hospital = Dict[str, object]

URBAN_CENTRES = {
  "harare",
  "bulawayo",
  "gweru",
  "mutare",
  "masvingo",
  "kwekwe",
  "chitungwiza",
  "queque",
  "chinhoyi",
  "bindura",
  "victoria falls",
  "vic falls",
}

TIER1_SPECIALISTS = {
  "oncology",
  "cardiology",
  "neurosurgery",
  "icu",
  "critical care",
  "trauma",
  "hematology",
  "neonatology",
}


def normalize_text(value: str) -> str:
  """Lowercase, strip punctuation, and collapse whitespace for matching."""
  cleaned = re.sub(r"[^a-z0-9]+", " ", (value or "").lower())
  return re.sub(r"\s+", " ", cleaned).strip()


def make_key(name: str, district: str, province: str = "") -> str:
  return f"{normalize_text(name)}::{normalize_text(district)}::{normalize_text(province)}"


def slugify(name: str, district: str) -> str:
  slug = re.sub(r"[^a-z0-9]+", "-", f"{name}-{district}".lower()).strip("-")
  return slug or f"facility-{int(dt.datetime.now().timestamp())}"


def classify_facility_type(record: Hospital) -> str:
  """Classify a human-friendly facility type based on names and hints."""
  explicit = (record.get("facility_type") or "").strip()
  if explicit:
    return explicit

  name = normalize_text(str(record.get("name", "")))
  category = normalize_text(str(record.get("category", "")))
  type_hint = normalize_text(str(record.get("type", "")))

  for key, label in [
    ("central hospital", "Central Hospital"),
    ("provincial hospital", "Provincial Hospital"),
    ("district hospital", "District Hospital"),
    ("mission hospital", "Mission Hospital"),
    ("polyclinic", "Polyclinic"),
    ("private hospital", "Private Hospital"),
    ("clinic", "Clinic"),
    ("pharmacy", "Pharmacy"),
    ("optician", "Optician"),
    ("dental", "Dental Clinic"),
    ("laboratory", "Lab"),
  ]:
    if key in name or key in category or key in type_hint:
      return label

  if "mission" in type_hint or "church" in type_hint:
    return "Mission Hospital"
  if "hospital" in category or "hospital" in type_hint:
    return "Hospital"
  if "pharmacy" in category:
    return "Pharmacy"
  if "clinic" in category:
    return "Clinic"
  return "Health Facility"


def infer_rural_urban(record: Hospital) -> str:
  """Infer rural/urban flag using simple heuristics."""
  explicit = (record.get("rural_urban") or "").title()
  if explicit:
    return explicit

  district = normalize_text(str(record.get("district") or record.get("city") or ""))
  if district in URBAN_CENTRES:
    return "Urban"
  if "rural" in normalize_text(str(record.get("name", ""))):
    return "Rural"
  if "clinic" in normalize_text(str(record.get("category", ""))):
    return "Rural"
  return "Urban" if district else "Peri-urban"


def infer_default_services(record: Hospital) -> List[str]:
  """Infer likely services based on facility type and ownership.

  Only fills gaps; explicit services from sources are preserved.
  """
  facility_type = classify_facility_type(record)
  ownership = (record.get("ownership") or "").lower()
  services: List[str] = []

  if "Central" in facility_type or "Provincial" in facility_type:
    services = ["ER", "Maternity", "Theatre", "ICU", "Lab", "X-Ray", "Inpatient"]
  elif "District" in facility_type:
    services = ["ER", "Maternity", "Lab", "Inpatient"]
  elif "Clinic" in facility_type:
    services = ["OPD", "MCH", "Immunisation", "HIV"]
  elif "Pharmacy" in facility_type:
    services = ["Dispensary"]

  if "mission" in ownership and "Hospital" in facility_type:
    services = ["ER", "Maternity", "Lab", "Inpatient"]

  return services


def tier_from_record(record: Hospital) -> Optional[str]:
  """Apply MoHCC-style tiering with 2025 rules."""
  bed_count = record.get("bed_count") if isinstance(record.get("bed_count"), int) else None
  services = {normalize_text(s) for s in record.get("services", [])}
  type_value = normalize_text(str(record.get("facility_type", "")))

  has_tier1 = any(key in services for key in TIER1_SPECIALISTS) or bed_count is not None and bed_count >= 350
  is_central = "central" in type_value or "referral" in type_value or "teaching" in type_value
  if has_tier1 or is_central:
    return "Tier 1"

  if (bed_count and 120 <= bed_count <= 349) or "provincial" in type_value or "district" in type_value:
    return "Tier 2"

  return "Tier 3"


def clean_phone(value: Optional[str]) -> Optional[str]:
  if not value:
    return None
  return re.sub(r"\s+", " ", value).strip()


def open_hours_flag(record: Hospital) -> bool:
  hours = normalize_text(str(record.get("operating_hours") or ""))
  return "24" in hours or "24/7" in hours or "24 7" in hours


def load_json(path: pathlib.Path) -> List[Hospital]:
  with path.open() as fh:
    return json.load(fh)


def load_csv(path: pathlib.Path) -> List[Hospital]:
  rows: List[Hospital] = []
  with path.open(newline="") as fh:
    reader = csv.DictReader(fh)
    rows.extend(row for row in reader)
  return rows


def load_xlsx(path: pathlib.Path) -> List[Hospital]:
  """Load rows from an XLSX file when openpyxl is available."""

  try:
    import openpyxl  # type: ignore
  except ImportError:
    print(f"Skipping {path.name} (openpyxl not installed)")
    return []

  workbook = openpyxl.load_workbook(path)
  sheet = workbook.active
  headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(max_row=1))]
  facilities: List[Hospital] = []
  for row in sheet.iter_rows(min_row=2, values_only=True):
    record = {headers[idx]: value for idx, value in enumerate(row) if headers[idx]}
    facilities.append(record)
  return facilities


def load_pdf_tables(path: pathlib.Path) -> List[Hospital]:
  """Load tabular data from a PDF when pdfplumber is available."""

  try:
    import pdfplumber  # type: ignore
  except ImportError:
    print(f"Skipping {path.name} (pdfplumber not installed)")
    return []

  facilities: List[Hospital] = []
  with pdfplumber.open(path) as pdf:
    for page in pdf.pages:
      for table in page.extract_tables() or []:
        if not table or len(table) < 2:
          continue
        headers = [str(cell).strip() if cell else "" for cell in table[0]]
        for row in table[1:]:
          record: Hospital = {}
          for idx, cell in enumerate(row):
            header = headers[idx] if idx < len(headers) else f"column_{idx}"
            if header:
              record[header] = str(cell).strip() if cell else ""
          if record:
            facilities.append(record)
  return facilities


def load_html_tables(path: pathlib.Path) -> List[Hospital]:
  """Parse HTML tables into row dicts."""

  from bs4 import BeautifulSoup  # type: ignore

  facilities: List[Hospital] = []
  soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
  for table in soup.find_all("table"):
    headers: List[str] = []
    header_row = table.find("tr")
    if header_row:
      headers = [th.get_text(strip=True) for th in header_row.find_all(["th", "td"])]
    for row in table.find_all("tr")[1:]:
      cells = [cell.get_text(strip=True) for cell in row.find_all(["td", "th"])]
      if not cells:
        continue
      record: Hospital = {}
      for idx, cell in enumerate(cells):
        header = headers[idx] if idx < len(headers) and headers[idx] else f"column_{idx}"
        record[header] = cell
      facilities.append(record)
  return facilities


def coerce_bool(value: object) -> bool:
  if isinstance(value, bool):
    return value
  text = normalize_text(str(value))
  return text in {"true", "yes", "y", "1", "24", "24 7", "247", "24/7"}


def coerce_float(value: object) -> Optional[float]:
  try:
    return float(value)
  except (TypeError, ValueError):
    return None


def coerce_list(value: object) -> List[str]:
  if isinstance(value, list):
    return [str(item).strip() for item in value if str(item).strip()]
  if isinstance(value, str):
    parts = re.split(r"[,;/]|\s{2,}", value)
    return [part.strip() for part in parts if part.strip()]
  return []


def normalize_raw_record(record: Hospital, source_label: str = "") -> Hospital:
  """Coerce loose raw fields and tag their source label for provenance."""

  normalised: Hospital = dict(record)
  field_aliases = {
    "provider": "name",
    "service provider": "name",
    "premises name": "name",
    "provider name": "name",
    "town": "city",
    "city/town": "city",
    "location": "city",
    "province/state": "province",
    "tel": "phone",
    "telephone": "phone",
  }
  for alias, target in field_aliases.items():
    if alias in normalised and target not in normalised:
      normalised[target] = normalised.pop(alias)
  if source_label and not normalised.get("source"):
    normalised["source"] = [source_label]
  elif isinstance(normalised.get("source"), str):
    normalised["source"] = [normalised["source"]]

  normalised["services"] = coerce_list(normalised.get("services") or normalised.get("specialists"))
  normalised["medical_aids"] = coerce_list(normalised.get("medical_aids") or normalised.get("accepted_payments"))
  normalised["open_24h"] = coerce_bool(normalised.get("open_24h") or normalised.get("open_hrs"))
  normalised["lat"] = normalised.get("lat") or coerce_float(normalised.get("latitude"))
  normalised["lon"] = normalised.get("lon") or coerce_float(normalised.get("longitude"))
  normalised["confidence"] = normalised.get("confidence") or "medium"
  return normalised


def load_raw_sources() -> List[Hospital]:
  facilities: List[Hospital] = []
  fetch_remote_sources()
  if not RAW_DIR.exists():
    return facilities

  for file in RAW_DIR.glob("*.*"):
    source_label = file.stem
    raw_records: List[Hospital] = []
    if file.suffix.lower() == ".json":
      raw_records = load_json(file)
    elif file.suffix.lower() == ".csv":
      raw_records = load_csv(file)
    elif file.suffix.lower() in {".xlsx", ".xls"}:
      raw_records = load_xlsx(file)
    elif file.suffix.lower() == ".pdf":
      raw_records = load_pdf_tables(file)
    elif file.suffix.lower() in {".htm", ".html"}:
      raw_records = load_html_tables(file)

    for record in raw_records:
      facilities.append(normalize_raw_record(record, source_label))
  return facilities


def scraper_ministry_portal() -> List[Hospital]:
  """TODO: real MoHCC ingestion. Currently seeds known provincial hospitals."""
  return [
    {
      "name": "Gweru Provincial Hospital",
      "province": "Midlands",
      "city": "Gweru",
      "district": "Gweru",
      "address": "Hospital Rd, Gweru",
      "type": "public",
      "ownership": "Government",
      "bed_count": 320,
      "services": ["ER", "Maternity"],
      "phone": "+263-54-222-333",
      "website": "",
      "facility_type": "Provincial Hospital",
      "operating_hours": "24/7 for emergencies; outpatient 08:00-17:00",
    },
  ]


def scraper_private_networks() -> List[Hospital]:
  return [
    {
      "name": "Borrowdale Trauma Centre",
      "province": "Harare",
      "city": "Harare",
      "district": "Harare",
      "address": "Borrowdale Rd, Harare",
      "type": "private",
      "ownership": "Private",
      "bed_count": 80,
      "services": ["Trauma", "ICU"],
      "phone": "+263-4-870-000",
      "website": "https://www.traumacentre.co.zw",
      "facility_type": "Private Hospital",
      "operating_hours": "24/7",
    },
  ]


def scraper_google_places_stub() -> List[Hospital]:
  """Placeholder for Google-sourced updates; currently seeds Hwange/Vic Falls."""
  return [
    {
      "name": "Hwange Colliery Hospital",
      "province": "Matabeleland North",
      "city": "Hwange",
      "district": "Hwange",
      "address": "Lusumbami, Hwange",
      "ownership": "Corporate",
      "facility_type": "District Hospital",
      "services": ["ER", "Maternity", "Lab"],
      "phone": "+263 281 214 1234",
      "latitude": -18.364,
      "longitude": 26.501,
      "source": "google_stub",
    },
    {
      "name": "Victoria Falls Hospital",
      "province": "Matabeleland North",
      "city": "Victoria Falls",
      "district": "Victoria Falls",
      "address": "Park Way, Victoria Falls",
      "ownership": "Government",
      "facility_type": "District Hospital",
      "services": ["ER", "Maternity", "Lab"],
      "phone": "+263 213 284 3216",
      "latitude": -17.926,
      "longitude": 25.842,
      "source": "google_stub",
    },
  ]


SCRAPERS = [
  scraper_ministry_portal,
  scraper_private_networks,
  scraper_google_places_stub,
]


def merge_field(primary: object, secondary: object) -> object:
  return primary if primary not in (None, "", []) else secondary


def merge_sources(a: Iterable[str], b: Iterable[str]) -> List[str]:
  merged = {s for s in a if s} | {s for s in b if s}
  return sorted(merged) if merged else []


def deduplicate_facilities(facilities: List[Hospital]) -> List[Hospital]:
  """Merge near-duplicate facilities using fuzzy name matching.

  Facilities are compared within the same province/district context. The best
  canonical name wins; alternate spellings are captured in ``aliases``.
  """
  canonical: List[Hospital] = []
  for record in facilities:
    district = record.get("district") or record.get("city") or ""
    province = record.get("province") or ""
    key = make_key(record.get("name", ""), district, province)
    matched: Optional[Hospital] = None
    matched_score = 0
    for existing in canonical:
      same_province = normalize_text(existing.get("province", "")) == normalize_text(province)
      same_district = normalize_text(existing.get("district", "")) == normalize_text(district)
      if not (same_province or same_district):
        continue
      score = SequenceMatcher(
        None,
        normalize_text(existing.get("name", "")),
        normalize_text(record.get("name", "")),
      ).ratio() * 100
      if score > 88 and score > matched_score:
        matched = existing
        matched_score = score
    if not matched:
      record.setdefault("aliases", [])
      record.setdefault("source", [])
      record.setdefault("confidence", "medium")
      if not record.get("verified"):
        record["verified"] = any(src in TRUSTED_SOURCES for src in record.get("source", []))
      record["_key"] = key
      canonical.append(record)
      continue

    aliases = set(matched.get("aliases", [])) | {record.get("name", "")}
    matched["aliases"] = sorted({a for a in aliases if a})
    matched_sources = merge_sources(matched.get("source", []), record.get("source", []))
    matched["source"] = matched_sources
    matched["confidence"] = "high" if matched_score > 92 else matched.get("confidence", "medium")
    matched["verified"] = matched.get("verified") or any(src in TRUSTED_SOURCES for src in matched_sources)

    for field in [
      "facility_type",
      "ownership",
      "rural_urban",
      "province",
      "district",
      "ward",
      "city",
      "address",
      "emergency_level",
      "cost_band",
      "tier",
      "website",
      "email",
      "last_verified",
    ]:
      matched[field] = merge_field(matched.get(field), record.get(field))

    matched_services = set(matched.get("services", []) or []) | set(record.get("services", []) or [])
    matched["services"] = sorted(matched_services) if matched_services else []

    matched_aids = set(matched.get("medical_aids", []) or []) | set(record.get("medical_aids", []) or [])
    matched["medical_aids"] = sorted(matched_aids) if matched_aids else []

    matched_aliases = set(matched.get("aliases", []) or []) | set(record.get("aliases", []) or [])
    matched["aliases"] = sorted(a for a in matched_aliases if a)

    for coord_field in ["lat", "lon", "latitude", "longitude"]:
      if matched.get("lat") and matched.get("lon"):
        break
      if coord_field in record:
        matched["lat"] = record.get("lat") or record.get("latitude")
        matched["lon"] = record.get("lon") or record.get("longitude")

    for phone_field in ["phone", "whatsapp"]:
      matched[phone_field] = merge_field(matched.get(phone_field), record.get(phone_field))
  return canonical


def map_to_schema(record: Hospital) -> Hospital:
  """Transform a heterogeneous raw record into the export schema."""
  facility_type = classify_facility_type(record)
  rural_urban = infer_rural_urban(record)
  services = record.get("services") or record.get("specialists") or []
  if not services:
    services = infer_default_services({"facility_type": facility_type, "ownership": record.get("ownership")})
  else:
    services = list(dict.fromkeys(s.strip() for s in services if s))

  services = services or infer_default_services({"facility_type": facility_type, "ownership": record.get("ownership")})

  tier_value = None
  tier_raw = str(record.get("tier") or "").strip()
  if tier_raw:
    normalized = tier_raw.lower().replace("tier", "").replace(" ", "")
    normalized = normalized[1:] if normalized.startswith("t") else normalized
    if normalized.isdigit():
      tier_value = f"Tier {normalized}"
    elif tier_raw in {"Tier 1", "Tier 2", "Tier 3"}:
      tier_value = tier_raw

  sources = record.get("source") if isinstance(record.get("source"), list) else ([record.get("source")] if record.get("source") else [])

  export: Hospital = {
    "id": record.get("id") or slugify(record.get("name", "facility"), record.get("district") or record.get("city") or "zw"),
    "name": record.get("name", "Unnamed Facility"),
    "aliases": record.get("aliases", []),
    "facility_type": facility_type,
    "ownership": (record.get("ownership") or record.get("type") or "").title() or None,
    "rural_urban": rural_urban,
    "province": record.get("province") or "",
    "district": record.get("district") or record.get("city") or "",
    "ward": record.get("ward") or "",
    "city": record.get("city") or record.get("district") or "",
    "address": record.get("address") or "",
    "services": services,
    "open_24h": bool(record.get("open_24h")) or open_hours_flag(record) or "central" in normalize_text(facility_type),
    "emergency_level": record.get("emergency_level") or ("Full" if "Hospital" in facility_type else "Basic"),
    "cost_band": record.get("cost_band") or None,
    "medical_aids": record.get("medical_aids") or record.get("accepted_payments") or [],
    "phone": clean_phone(record.get("phone")),
    "whatsapp": clean_phone(record.get("whatsapp") or record.get("phone")),
    "email": record.get("email") or None,
    "lat": record.get("lat") or record.get("latitude"),
    "lon": record.get("lon") or record.get("longitude"),
    "tier": tier_value or tier_from_record({"facility_type": facility_type, "bed_count": record.get("bed_count"), "services": services}),
    "last_verified": record.get("last_verified") or TODAY,
    "source": sources,
    "confidence": record.get("confidence") or "medium",
    "verified": record.get("verified") or any(src in TRUSTED_SOURCES for src in sources),
    "website": record.get("website") or "",
  }
  return export


def validate_facilities(facilities: List[Hospital]) -> List[Hospital]:
  cleaned: List[Hospital] = []
  for item in facilities:
    if not item.get("name") or not item.get("province"):
      continue
    cleaned.append(item)
  return cleaned


def run_pipeline() -> List[Hospital]:
  raw_records: List[Hospital] = []
  raw_records.extend(load_raw_sources())
  for scraper in SCRAPERS:
    raw_records.extend(scraper())

  deduped = deduplicate_facilities(raw_records)
  normalized = [map_to_schema(record) for record in deduped]
  validated = validate_facilities(normalized)

  for record in validated:
    if record.get("source"):
      record["confidence"] = "high" if len(record["source"]) > 1 else record.get("confidence", "medium")
    if record.get("open_24h") and record.get("emergency_level") == "Basic" and "Hospital" in record.get("facility_type", ""):
      record["emergency_level"] = "Full"

  validated.sort(key=lambda h: (h.get("province", ""), h.get("district", ""), h.get("name", "")))
  return validated


def save_records(records: List[Hospital]) -> None:
  SCRAPED_OUTPUT.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n")
  full_path = SCRAPED_OUTPUT.with_name("hospitals_scraped_full.json")
  full_path.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n")


def main() -> None:
  records = run_pipeline()
  save_records(records)
  print(f"Wrote {len(records)} facilities to {SCRAPED_OUTPUT}")


if __name__ == "__main__":
  main()
